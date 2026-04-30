"""
Pipeline K-Means + Gaussian Naive Bayes — Streamlit App Lengkap
Versi 2.0 — Excel identik dengan skrip standalone
"""

import streamlit as st
import pandas as pd
import numpy as np
import math, io, pickle, warnings
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import plotly.express as px
from datetime import datetime

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
# KONFIGURASI HALAMAN
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Pipeline K-Means + Naive Bayes",
    layout="wide",
    page_icon="🔬"
)

# ─────────────────────────────────────────────
# KONSTANTA WARNA EXCEL
# ─────────────────────────────────────────────
BIRU_TUA  = "1F4E79"; BIRU_MID  = "2E75B6"; BIRU_MUDA = "BDD7EE"
HIJAU_TUA = "375623"; HIJAU_MID = "70AD47"; HIJAU_MUDA= "E2EFDA"
MERAH_MUDA= "FCE4D6"; MERAH_TUA = "9C0006"; KUNING    = "FFF2CC"
ABU       = "D9D9D9"; PUTIH     = "FFFFFF"; HITAM     = "000000"
C1_BG="C6EFCE"; C2_BG="FFEB9C"; C3_BG="FFC7CE"
C1_FG="375623"; C2_FG="9C6500"; C3_FG="9C0006"

KET    = ["Mudah","Sedang","Sulit"]
LABEL  = ["C1 (Mudah)","C2 (Sedang)","C3 (Sulit)"]
BG_C   = [C1_BG, C2_BG, C3_BG]
FG_C   = [C1_FG, C2_FG, C3_FG]
COLORS_K  = ['#2E75B6','#70AD47','#E00000']
MARKERS_K = ['o','s','^']

# ─────────────────────────────────────────────
# HELPER FUNGSI STYLING EXCEL
# ─────────────────────────────────────────────
def solid(h):
    return PatternFill("solid", start_color=h, fgColor=h)

def aln(h="center", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def bdr():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, r, c, v, bold=False, bg=None, fg=HITAM, h="center",
       wrap=False, sz=11, border=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color=fg, size=sz, name="Calibri")
    if bg: cell.fill = solid(bg)
    cell.alignment = aln(h, wrap)
    if border: cell.border = bdr()
    return cell

def mg(ws, r1, c1, r2, c2, v, bold=False, bg=None, fg=HITAM, h="center",
       wrap=False, sz=11, italic=False):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=v)
    cell.font = Font(bold=bold, color=fg, size=sz, italic=italic, name="Calibri")
    if bg: cell.fill = solid(bg)
    cell.alignment = aln(h, wrap)
    cell.border = bdr()
    return cell


# ─────────────────────────────────────────────
# STEP 1: DETEKSI & LOAD DATA
# ─────────────────────────────────────────────
def load_and_prepare(df_raw):
    """
    Deteksi kolom secara fleksibel:
    - Nomor Soal
    - Persentase  ATAU  (Jumlah Benar + Total Siswa)
    - Waktu
    Jika persentase tidak ada, hitung dari jumlah_benar / total_siswa × 100
    """
    col_map = {}
    for col in df_raw.columns:
        c = str(col).lower().strip()
        # Nomor Soal
        if any(x in c for x in ['nomor soal','no soal','soal','no.','no ']):
            if 'Soal' not in col_map:
                col_map['Soal'] = col
        # Persentase
        if any(x in c for x in ['persentase','persen','p (%)', 'p(%)', '(%)']):
            col_map['Persentase'] = col
        # Waktu
        if any(x in c for x in ['waktu','time','detik','rata-rata waktu']):
            col_map['Waktu'] = col
        # Jumlah Benar
        if any(x in c for x in ['jumlah benar','jml benar','benar','correct','jwb benar']):
            col_map['JumlahBenar'] = col
        # Total Siswa
        if any(x in c for x in ['total siswa','jumlah siswa','jml siswa','siswa','peserta']):
            col_map['TotalSiswa'] = col

    # Validasi Soal
    if 'Soal' not in col_map:
        # coba kolom pertama sebagai soal
        col_map['Soal'] = df_raw.columns[0]

    # Hitung persentase jika tidak ada kolom langsung
    if 'Persentase' not in col_map:
        if 'JumlahBenar' in col_map and 'TotalSiswa' in col_map:
            df_raw = df_raw.copy()
            df_raw['_Persentase_calc'] = (
                pd.to_numeric(df_raw[col_map['JumlahBenar']], errors='coerce') /
                pd.to_numeric(df_raw[col_map['TotalSiswa']], errors='coerce') * 100
            ).round(2)
            col_map['Persentase'] = '_Persentase_calc'
        else:
            raise ValueError(
                "Kolom 'Persentase' tidak ditemukan dan tidak ada "
                "'Jumlah Benar' + 'Total Siswa' untuk menghitungnya.\n"
                f"Kolom tersedia: {list(df_raw.columns)}"
            )

    if 'Waktu' not in col_map:
        raise ValueError(
            f"Kolom 'Waktu' tidak ditemukan.\n"
            f"Kolom tersedia: {list(df_raw.columns)}"
        )

    # Ambil kolom yang diperlukan
    cols_needed = [col_map['Soal'], col_map['Persentase'], col_map['Waktu']]
    # Tambahkan JumlahBenar dan TotalSiswa jika ada
    extra_cols = []
    for key in ['JumlahBenar', 'TotalSiswa']:
        if key in col_map and col_map[key] in df_raw.columns:
            extra_cols.append(col_map[key])

    df = df_raw[cols_needed + extra_cols].copy().dropna(
        subset=[col_map['Soal'], col_map['Persentase'], col_map['Waktu']]
    ).reset_index(drop=True)

    rename = {
        col_map['Soal']       : 'Soal',
        col_map['Persentase'] : 'Persentase',
        col_map['Waktu']      : 'Waktu',
    }
    if 'JumlahBenar' in col_map and col_map['JumlahBenar'] in df.columns:
        rename[col_map['JumlahBenar']] = 'JumlahBenar'
    if 'TotalSiswa' in col_map and col_map['TotalSiswa'] in df.columns:
        rename[col_map['TotalSiswa']] = 'TotalSiswa'

    df = df.rename(columns=rename)
    df['Persentase'] = pd.to_numeric(df['Persentase'], errors='coerce').round(2)
    df['Waktu']      = pd.to_numeric(df['Waktu'],      errors='coerce').round(2)
    df = df.dropna(subset=['Persentase','Waktu']).reset_index(drop=True)

    # Normalisasi nama soal → S1, S2, ...
    soal_list = []
    for i, v in enumerate(df['Soal'], 1):
        s = str(v).strip()
        soal_list.append(
            f"S{int(s)}" if s.isdigit() else
            s.upper()   if s.lower().startswith('s') and s[1:].isdigit() else
            f"S{i}"
        )
    df['Soal'] = soal_list

    info = {
        'has_jumlah_benar': 'JumlahBenar' in df.columns,
        'has_total_siswa' : 'TotalSiswa'  in df.columns,
        'col_map'         : col_map,
    }
    return df, info


# ─────────────────────────────────────────────
# STEP 2: K-MEANS MANUAL
# ─────────────────────────────────────────────
def euclidean(p1, p2):
    return math.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)

def choose_centroids(data):
    df = data.copy()
    c1r = df.sort_values(['Persentase','Waktu'], ascending=[False,True]).iloc[0]
    c3r = df.sort_values(['Persentase','Waktu'], ascending=[True,False]).iloc[0]
    df['_dp'] = abs(df['Persentase'] - df['Persentase'].median())
    df['_dw'] = abs(df['Waktu']      - df['Waktu'].median())
    c2r = df.sort_values(['_dp','_dw']).iloc[0]
    return [
        (round(c1r['Persentase'],2), round(c1r['Waktu'],2)),
        (round(c2r['Persentase'],2), round(c2r['Waktu'],2)),
        (round(c3r['Persentase'],2), round(c3r['Waktu'],2)),
    ]

def run_kmeans(data):
    points    = list(zip(data['Persentase'], data['Waktu']))
    centroids = choose_centroids(data)
    history   = []

    for it in range(20):
        assignments = []
        dists       = []
        for p in points:
            d = [round(euclidean(p, centroids[k]), 2) for k in range(3)]
            assignments.append(d.index(min(d)))
            dists.append(d)

        new_cen = []
        for k in range(3):
            mems = [points[i] for i in range(len(points)) if assignments[i]==k]
            if mems:
                new_cen.append((
                    round(sum(m[0] for m in mems)/len(mems), 2),
                    round(sum(m[1] for m in mems)/len(mems), 2),
                ))
            else:
                new_cen.append(centroids[k])

        conv = (new_cen == centroids)
        history.append({
            'centroids'    : list(centroids),
            'assignments'  : assignments,
            'distances'    : dists,
            'new_centroids': new_cen,
            'converged'    : conv,
        })
        if conv:
            break
        centroids = new_cen

    return history


# ─────────────────────────────────────────────
# STEP 3: STRATIFIED SPLIT
# ─────────────────────────────────────────────
def stratified_split(data, test_ratio=0.25, random_state=42):
    np.random.seed(random_state)
    groups = defaultdict(list)
    for i, row in data.iterrows():
        groups[row['Keterangan']].append(i)

    train_idx, test_idx = [], []
    detail = {}
    for label in ['Mudah','Sedang','Sulit']:
        idxs = groups[label]
        if not idxs:
            detail[label] = {'total':0,'train':0,'test':0}
            continue
        np.random.shuffle(idxs)
        n_test  = max(1, round(len(idxs)*test_ratio))
        n_train = len(idxs) - n_test
        train_idx += idxs[:n_train]
        test_idx  += idxs[n_train:]
        detail[label] = {'total':len(idxs),'train':n_train,'test':n_test}

    train = data.loc[train_idx].reset_index(drop=True)
    test  = data.loc[test_idx].reset_index(drop=True)
    return train, test, detail


# ─────────────────────────────────────────────
# STEP 4: GAUSSIAN NAIVE BAYES MANUAL
# ─────────────────────────────────────────────
class GaussianNaiveBayes:
    def fit(self, X, y):
        self.classes_ = np.unique(y)
        self.priors_, self.means_, self.stds_ = {}, {}, {}
        for c in self.classes_:
            Xc = X[y == c]
            self.priors_[c] = len(Xc) / len(y)
            self.means_[c]  = Xc.mean(axis=0)
            self.stds_[c]   = Xc.std(axis=0) + 1e-9

    def _gaussian(self, x, mean, std):
        return (1/(np.sqrt(2*np.pi)*std)) * np.exp(-((x-mean)**2)/(2*std**2))

    def _log_posterior(self, x):
        res = {}
        for c in self.classes_:
            res[c] = (np.log(self.priors_[c]) +
                      np.sum(np.log(self._gaussian(x, self.means_[c], self.stds_[c]))))
        return res

    def predict(self, X):
        return np.array([max(self._log_posterior(x), key=self._log_posterior(x).get)
                         for x in X])

    def predict_proba(self, X):
        hasil = []
        for x in X:
            lp  = self._log_posterior(x)
            arr = np.array([lp[c] for c in self.classes_], dtype=float)
            arr -= arr.max()
            p    = np.exp(arr); p /= p.sum()
            hasil.append(p)
        return np.array(hasil)

    def save_bytes(self):
        buf = io.BytesIO()
        pickle.dump(self, buf)
        buf.seek(0)
        return buf


# ─────────────────────────────────────────────
# STEP 5: EVALUASI
# ─────────────────────────────────────────────
def cm_manual(y_true, y_pred, classes):
    cm = pd.DataFrame(0, index=classes, columns=classes)
    for t, p in zip(y_true, y_pred):
        cm.loc[t, p] += 1
    return cm

def metrics_manual(cm):
    hasil = {}
    for c in cm.index:
        tp = cm.loc[c,c]
        fp = cm[c].sum() - tp
        fn = cm.loc[c].sum() - tp
        pr = tp/(tp+fp) if (tp+fp)>0 else 0.0
        rc = tp/(tp+fn) if (tp+fn)>0 else 0.0
        f1 = 2*pr*rc/(pr+rc) if (pr+rc)>0 else 0.0
        hasil[c] = [pr, rc, f1]
    return pd.DataFrame(hasil, index=["Precision","Recall","F1"]).T


# ─────────────────────────────────────────────
# STEP 6: FUNGSI GRAFIK
# ─────────────────────────────────────────────
def make_kmeans_scatter(it_num, it_data, points, soal_list, converged=False):
    fig, ax = plt.subplots(figsize=(11,7))
    ax.set_facecolor('#F8F9FA'); fig.patch.set_facecolor('white')
    cents, assigns = it_data['centroids'], it_data['assignments']
    for k in range(3):
        xs = [points[i][0] for i in range(len(points)) if assigns[i]==k]
        ys = [points[i][1] for i in range(len(points)) if assigns[i]==k]
        ax.scatter(xs, ys, c=COLORS_K[k], marker=MARKERS_K[k], s=120,
                   zorder=3, alpha=0.85, edgecolors='white', linewidths=0.8)
        for i in range(len(points)):
            if assigns[i]==k:
                ax.annotate(soal_list[i], (points[i][0], points[i][1]),
                            textcoords="offset points", xytext=(5,3),
                            fontsize=8, color=COLORS_K[k], fontweight='bold')
    for k in range(3):
        ax.scatter(cents[k][0], cents[k][1], c=COLORS_K[k], marker='*',
                   s=500, zorder=5, edgecolors='black', linewidths=1.2)
        ax.add_patch(plt.Circle((cents[k][0],cents[k][1]),2.5,
                                color=COLORS_K[k],fill=False,
                                linestyle='--',linewidth=1,alpha=0.5))
    sfx = ' (KONVERGEN)' if converged else ''
    ax.set_xlabel('Persentase Jawaban Benar (%)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Waktu Rata-rata (detik)', fontsize=11, fontweight='bold')
    ax.set_title(f'Iterasi {it_num} — K-Means Clustering{sfx}\n'
                 f'C1={cents[0]}  C2={cents[1]}  C3={cents[2]}',
                 fontsize=12, fontweight='bold', pad=12)
    patches = [mpatches.Patch(color=COLORS_K[k], label=f'C{k+1} - {KET[k]}') for k in range(3)]
    patches.append(mpatches.Patch(color='gray', label='* = Centroid'))
    ax.legend(handles=patches, loc='upper right', fontsize=9, framealpha=0.9)
    ax.grid(True, linestyle='--', alpha=0.4, color='#CCCCCC')
    pcts = [p[0] for p in points]; wkts = [p[1] for p in points]
    pad_x = (max(pcts)-min(pcts))*0.15+3
    pad_y = (max(wkts)-min(wkts))*0.15+5
    ax.set_xlim(min(pcts)-pad_x, max(pcts)+pad_x+10)
    ax.set_ylim(min(wkts)-pad_y, max(wkts)+pad_y)
    plt.tight_layout(pad=1.5)
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(); buf.seek(0)
    return buf

def make_cm_heatmap(cm_df, classes):
    fig, ax = plt.subplots(figsize=(5,4))
    data_arr = cm_df.values.astype(float)
    im = ax.imshow(data_arr, cmap='Blues')
    ax.set_xticks(range(len(classes))); ax.set_yticks(range(len(classes)))
    ax.set_xticklabels(classes, fontsize=11, fontweight='bold')
    ax.set_yticklabels(classes, fontsize=11, fontweight='bold')
    ax.set_xlabel('Prediksi', fontsize=12, fontweight='bold')
    ax.set_ylabel('Aktual',   fontsize=12, fontweight='bold')
    ax.set_title('Confusion Matrix\nGaussian Naive Bayes',
                 fontsize=13, fontweight='bold', pad=10)
    for i in range(len(classes)):
        for j in range(len(classes)):
            val   = int(data_arr[i,j])
            color = 'white' if data_arr[i,j]>data_arr.max()*0.6 else 'black'
            ax.text(j, i, str(val), ha='center', va='center',
                    fontsize=16, fontweight='bold', color=color)
    plt.colorbar(im, ax=ax)
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close(); buf.seek(0)
    return buf

def make_metrics_bar(met_df, acc):
    fig, ax = plt.subplots(figsize=(7,4.5))
    classes = met_df.index.tolist()
    x = np.arange(len(classes)); w = 0.25
    colors_bar = ['#2E75B6','#70AD47','#ED7D31']
    for i, (col, clr) in enumerate(zip(['Precision','Recall','F1'], colors_bar)):
        vals = [met_df.loc[c, col] for c in classes]
        bars = ax.bar(x+i*w, vals, w, label=col, color=clr, alpha=0.85, edgecolor='white')
        for bar, val in zip(bars, vals):
            ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.01,
                    f'{val:.2f}', ha='center', va='bottom', fontsize=9, fontweight='bold')
    ax.set_xticks(x+w); ax.set_xticklabels(classes, fontsize=11, fontweight='bold')
    ax.set_ylim(0, 1.15)
    ax.set_ylabel('Score', fontsize=11, fontweight='bold')
    ax.set_title(f'Precision / Recall / F1-Score per Kelas\nAkurasi = {acc*100:.2f}%',
                 fontsize=12, fontweight='bold', pad=10)
    ax.axhline(y=acc, color='red', linestyle='--', linewidth=1.5,
               label=f'Akurasi={acc*100:.2f}%')
    ax.legend(fontsize=10, framealpha=0.9)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close(); buf.seek(0)
    return buf

def make_scatter_nb(train, test, y_pred_test, classes):
    fig, ax = plt.subplots(figsize=(8,5.5))
    ax.set_facecolor('#F8F9FA'); fig.patch.set_facecolor('white')
    color_map = {'Mudah':COLORS_K[0],'Sedang':COLORS_K[1],'Sulit':COLORS_K[2]}
    for c in classes:
        sub = train[train['Keterangan']==c]
        ax.scatter(sub['Persentase'], sub['Waktu'], c=color_map.get(c,'gray'),
                   marker='o', s=80, alpha=0.6, edgecolors='white',
                   linewidths=0.5, label=f'Train - {c}')
    for (_, row), pred in zip(test.iterrows(), y_pred_test):
        c       = row['Keterangan']
        correct = (pred == c)
        ax.scatter(row['Persentase'], row['Waktu'], c=color_map.get(c,'gray'),
                   marker='D', s=120,
                   edgecolors='black' if correct else 'red', linewidths=2, zorder=5)
        ax.annotate(row['Soal'], (row['Persentase'], row['Waktu']),
                    textcoords="offset points", xytext=(5,3),
                    fontsize=7, color='black', fontweight='bold')
    ax.set_xlabel('Persentase Jawaban Benar (%)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Waktu Rata-rata (detik)',       fontsize=11, fontweight='bold')
    ax.set_title('Sebaran Data Train vs Test — Gaussian Naive Bayes\n'
                 'Lingkaran=Train  |  Berlian=Test  |  Border Merah=Prediksi Salah',
                 fontsize=11, fontweight='bold', pad=10)
    patches  = [mpatches.Patch(color=color_map.get(c,'gray'), label=c) for c in classes]
    patches += [mpatches.Patch(facecolor='white', edgecolor='black', label='Test Benar'),
                mpatches.Patch(facecolor='white', edgecolor='red',   label='Test Salah')]
    ax.legend(handles=patches, loc='upper right', fontsize=9, framealpha=0.9)
    ax.grid(True, linestyle='--', alpha=0.4)
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close(); buf.seek(0)
    return buf

def make_gaussian_pdf(model, train, classes):
    fig, axes = plt.subplots(1, 2, figsize=(12,5))
    fig.patch.set_facecolor('white')
    feature_names = ['Persentase (%)', 'Waktu (detik)']
    colors_pdf = ['#2E75B6','#70AD47','#E00000']
    linestyles = ['-','--','-.']
    for fi, (ax, fname) in enumerate(zip(axes, feature_names)):
        ax.set_facecolor('#F8F9FA')
        all_vals = train[['Persentase','Waktu']].values[:, fi]
        rng  = all_vals.max() - all_vals.min()
        x    = np.linspace(all_vals.min()-rng*0.2, all_vals.max()+rng*0.2, 300)
        for k, c in enumerate(classes):
            mean = model.means_[c][fi]; std = model.stds_[c][fi]
            pdf  = (1/(np.sqrt(2*np.pi)*std)) * np.exp(-((x-mean)**2)/(2*std**2))
            ax.plot(x, pdf, color=colors_pdf[k], linewidth=2.5,
                    linestyle=linestyles[k], label=f'{c} (μ={mean:.2f}, σ={std:.2f})')
            ax.axvline(mean, color=colors_pdf[k], linewidth=1, linestyle=':', alpha=0.6)
            ax.fill_between(x, pdf, alpha=0.08, color=colors_pdf[k])
            vals = train[train['Keterangan']==c][['Persentase','Waktu']].values[:, fi]
            ax.scatter(vals, np.zeros(len(vals))-0.002,
                       color=colors_pdf[k], marker='|', s=60, alpha=0.7, linewidths=1.5)
        ax.set_xlabel(fname, fontsize=12, fontweight='bold')
        ax.set_ylabel('Densitas Probabilitas', fontsize=11)
        ax.set_title(f'Distribusi Gaussian — {fname}', fontsize=12, fontweight='bold', pad=10)
        ax.legend(fontsize=9, framealpha=0.9)
        ax.grid(True, linestyle='--', alpha=0.4); ax.set_ylim(bottom=-0.005)
    fig.suptitle('Kurva Gaussian PDF per Kelas — Gaussian Naive Bayes',
                 fontsize=14, fontweight='bold', y=1.01)
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close(); buf.seek(0)
    return buf

def make_proba_bar(test, y_pred, y_proba, classes):
    n = len(test); soal = list(test['Soal'])
    colors_p = ['#2E75B6','#70AD47','#E00000']
    fig, ax = plt.subplots(figsize=(10, max(5, n*0.55)))
    fig.patch.set_facecolor('white'); ax.set_facecolor('#F8F9FA')
    y_pos = np.arange(n); lefts = np.zeros(n)
    for k, c in enumerate(classes):
        probs = [y_proba[i][k]*100 for i in range(n)]
        bars  = ax.barh(y_pos, probs, left=lefts, color=colors_p[k],
                        alpha=0.85, label=c, edgecolor='white', linewidth=0.5)
        for i, (bar, p) in enumerate(zip(bars, probs)):
            if p > 8:
                ax.text(lefts[i]+p/2, i, f'{p:.1f}%',
                        ha='center', va='center', fontsize=8,
                        fontweight='bold', color='white')
        lefts += np.array(probs)
    for i, (s, pred, actual) in enumerate(zip(soal, y_pred, test['Keterangan'])):
        correct = (pred == actual)
        ax.text(101, i,
                ('V ' if correct else 'X ') + s,
                va='center', fontsize=8.5, fontweight='bold',
                color='#375623' if correct else '#9C0006')
    ax.set_yticks(y_pos); ax.set_yticklabels(soal, fontsize=9)
    ax.set_xlim(0, 118)
    ax.set_xlabel('Probabilitas (%)', fontsize=11, fontweight='bold')
    ax.set_title('Probabilitas Prediksi per Soal Uji — Softmax Output',
                 fontsize=12, fontweight='bold', pad=10)
    ax.legend(loc='lower right', fontsize=10, framealpha=0.9)
    ax.axvline(50, color='gray', linewidth=0.8, linestyle='--', alpha=0.5)
    ax.grid(axis='x', linestyle='--', alpha=0.3); ax.invert_yaxis()
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close(); buf.seek(0)
    return buf

def make_pie_cluster(data, classes):
    counts = [list(data['Keterangan']).count(c) for c in classes]
    colors_pie = ['#2E75B6','#70AD47','#E00000']; explode = [0.04]*3
    fig, axes = plt.subplots(1, 2, figsize=(11,5))
    fig.patch.set_facecolor('white')
    ax = axes[0]
    wedges, _, autotexts = ax.pie(counts, autopct='%1.1f%%', colors=colors_pie,
                                   explode=explode, startangle=90, pctdistance=0.75,
                                   wedgeprops={'edgecolor':'white','linewidth':2})
    for at in autotexts:
        at.set_fontsize(11); at.set_fontweight('bold'); at.set_color('white')
    ax.set_title('Distribusi Cluster Hasil K-Means', fontsize=12, fontweight='bold', pad=12)
    ax.legend(wedges,
              [f'{c}: {n} soal ({n/sum(counts)*100:.1f}%)'
               for c, n in zip(classes, counts)],
              loc='lower center', bbox_to_anchor=(0.5,-0.12), fontsize=10, ncol=1)
    ax2 = axes[1]; ax2.set_facecolor('#F8F9FA')
    bars = ax2.bar(classes, counts, color=colors_pie, alpha=0.85,
                   edgecolor='white', linewidth=1.5, width=0.55)
    for bar, n in zip(bars, counts):
        ax2.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.3,
                 str(n), ha='center', va='bottom', fontsize=13, fontweight='bold',
                 color=bar.get_facecolor())
    ax2.set_ylabel('Jumlah Soal', fontsize=11, fontweight='bold')
    ax2.set_title('Jumlah Soal per Kelas Hasil K-Means', fontsize=12, fontweight='bold', pad=10)
    ax2.set_ylim(0, max(counts)*1.18); ax2.grid(axis='y', linestyle='--', alpha=0.4)
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close(); buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# STEP 7: WRITER SHEET EXCEL
# ─────────────────────────────────────────────
def write_sheet_data_asli(ws, df_raw):
    for c, col in enumerate(df_raw.columns, 1):
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(str(col))+2)
        sc(ws, 1, c, str(col), bold=True, bg=BIRU_TUA, fg=PUTIH)
    for r, row_data in enumerate(df_raw.itertuples(index=False), 2):
        bg = ABU if r%2==0 else PUTIH
        for c, val in enumerate(row_data, 1):
            v = "" if (isinstance(val, float) and math.isnan(val)) else val
            sc(ws, r, c, v, bg=bg)

def write_sheet_persentase(ws, data, has_jumlah_benar, has_total_siswa):
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['F'].width = 30

    headers = ['Soal']
    if has_jumlah_benar: headers.append('Jumlah Benar')
    if has_total_siswa:  headers.append('Total Siswa')
    headers.append('Persentase (%)')

    for c, h in enumerate(headers, 1):
        sc(ws, 1, c, h, bold=True, bg=BIRU_TUA, fg=PUTIH)

    # Info rumus di kolom F
    sc(ws, 1, 6, 'RUMUS', bold=True, bg=BIRU_MUDA, fg=BIRU_TUA)
    sc(ws, 2, 6, 'P = (Jumlah Benar / Total Siswa) x 100', bg=KUNING, h='left', wrap=True)

    for r, (_, row) in enumerate(data.iterrows(), 2):
        bg = ABU if r%2==0 else PUTIH
        col = 1
        sc(ws, r, col, row['Soal'], bg=bg); col += 1
        if has_jumlah_benar and 'JumlahBenar' in data.columns:
            sc(ws, r, col, row.get('JumlahBenar',''), bg=bg); col += 1
        if has_total_siswa and 'TotalSiswa' in data.columns:
            sc(ws, r, col, row.get('TotalSiswa',''), bg=bg); col += 1
        sc(ws, r, col, round(row['Persentase'], 2), bg=bg)

def write_sheet_waktu(ws, data):
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['E'].width = 38
    sc(ws, 1, 1, 'Soal',  bold=True, bg=BIRU_TUA, fg=PUTIH)
    sc(ws, 1, 2, 'Waktu (detik)', bold=True, bg=BIRU_TUA, fg=PUTIH)
    sc(ws, 1, 4, 'RUMUS', bold=True, bg=BIRU_MUDA, fg=BIRU_TUA)
    sc(ws, 2, 4, 'Waktu Rata-rata = (waktu tercepat + terlama) / 2',
       bg=KUNING, h='left', wrap=True)
    ws.row_dimensions[2].height = 32
    for r, (_, row) in enumerate(data.iterrows(), 2):
        bg = ABU if r%2==0 else PUTIH
        sc(ws, r, 1, row['Soal'], bg=bg)
        sc(ws, r, 2, int(row['Waktu']), bg=bg)

def write_iterasi_sheet(ws, it_num, it_data, data, prev_assigns=None):
    soal  = list(data['Soal']); pcts = list(data['Persentase']); wkts = list(data['Waktu'])
    cents = it_data['centroids']; assigns = it_data['assignments']
    dists = it_data['distances']; new_cen = it_data['new_centroids']
    converged = it_data['converged']
    for col, w in zip('ABCDEFGHI', [14,14,14,14,14,12,12,12,16]):
        ws.column_dimensions[col].width = w
    row = 1
    c_str = " | ".join(f"C{i+1}=({cents[i][0]}, {cents[i][1]})" for i in range(3))
    mg(ws,row,1,row,9, f"ITERASI {it_num} — Centroid: {c_str}",
       bold=True, bg=BIRU_TUA, fg=PUTIH, sz=12)
    ws.row_dimensions[row].height = 22; row += 2
    mg(ws,row,1,row,9,"CENTROID YANG DIGUNAKAN PADA ITERASI INI",
       bold=True, bg=BIRU_MUDA, fg=BIRU_TUA); row += 1
    for k in range(3):
        mg(ws,row,1,row,2,LABEL[k],bold=True,bg=BG_C[k],fg=FG_C[k])
        mg(ws,row,3,row,4,"",bg=BG_C[k])
        mg(ws,row,5,row,5,f"Persentase (y1) = {cents[k][0]}%",bg=BG_C[k],fg=FG_C[k])
        mg(ws,row,6,row,6,"",bg=BG_C[k])
        mg(ws,row,7,row,7,f"Waktu (y2) = {cents[k][1]} detik",bg=BG_C[k],fg=FG_C[k])
        mg(ws,row,8,row,9,"",bg=BG_C[k]); row += 1
    row += 1
    mg(ws,row,1,row,9,
       "d = SQRT((x1-y1)^2 + (x2-y2)^2)  |  x1=%soal, x2=waktu soal, y1=%centroid, y2=waktu centroid",
       italic=True, bg=KUNING, fg="7F7F7F"); row += 2
    for c, h in enumerate(["No\nSoal","Persentase\n(%)","Waktu\n(detik)",
                            "d(C1)\nMudah","d(C2)\nSedang","d(C3)\nSulit",
                            "Cluster","Keterangan","Ket\nWarna"],1):
        sc(ws,row,c,h,bold=True,bg=BIRU_TUA,fg=PUTIH,wrap=True)
    ws.row_dimensions[row].height = 32; row += 1
    for i in range(len(soal)):
        k = assigns[i]; bg = BG_C[k]; fg = FG_C[k]
        sl = soal[i] + (" <" if prev_assigns and prev_assigns[i]!=k else "")
        sc(ws,row,1,sl,bg=bg,fg=fg); sc(ws,row,2,round(pcts[i],2),bg=bg,fg=fg)
        sc(ws,row,3,int(wkts[i]),bg=bg,fg=fg)
        sc(ws,row,4,dists[i][0],bg=bg,fg=fg); sc(ws,row,5,dists[i][1],bg=bg,fg=fg)
        sc(ws,row,6,dists[i][2],bg=bg,fg=fg)
        sc(ws,row,7,f"C{k+1}",bold=True,bg=bg,fg=fg)
        sc(ws,row,8,KET[k],bold=True,bg=bg,fg=fg)
        sc(ws,row,9,"",bg=bg); row += 1
    row += 1
    mg(ws,row,1,row,9,"RINGKASAN CLUSTER",bold=True,bg=BIRU_MUDA,fg=BIRU_TUA); row += 1
    for k in range(3):
        mems = [soal[i] for i in range(len(soal)) if assigns[i]==k]
        mg(ws,row,1,row,2,f"C{k+1} — {KET[k]}: {len(mems)} soal",bold=True,bg=BG_C[k],fg=FG_C[k])
        mg(ws,row,3,row,9,", ".join(mems),bg=BG_C[k],fg=FG_C[k],h="left",wrap=True)
        ws.row_dimensions[row].height = 18; row += 1
    row += 1
    mg(ws,row,1,row,9,"PERHITUNGAN CENTROID BARU  (Ck = Sx / n)",
       bold=True, bg=BIRU_MUDA, fg=BIRU_TUA); row += 1
    for k in range(3):
        mp = [pcts[i] for i in range(len(soal)) if assigns[i]==k]
        mw = [wkts[i] for i in range(len(soal)) if assigns[i]==k]
        n  = len(mp); sp = round(sum(mp),2); sw = round(sum(mw),2)
        bg = BG_C[k]; fg = FG_C[k]
        mg(ws,row,1,row,2,LABEL[k],bold=True,bg=bg,fg=fg)
        mg(ws,row,3,row,4,f"n = {n}",bg=bg,fg=fg)
        mg(ws,row,5,row,5,f"S% = {sp}  |  SWaktu = {int(sw)}",bg=bg,fg=fg)
        mg(ws,row,6,row,6,"",bg=bg)
        mg(ws,row,7,row,7,f"Ck % = {sp}/{n} = {new_cen[k][0]}",bg=bg,fg=fg,wrap=True)
        mg(ws,row,8,row,8,"",bg=bg)
        mg(ws,row,9,row,9,f"Ck t = {int(sw)}/{n} = {new_cen[k][1]}",bg=bg,fg=fg,wrap=True)
        row += 1
    row += 1
    if converged:
        msg = "CENTROID TIDAK BERUBAH — K-Means KONVERGEN! Proses selesai."
        mbg = HIJAU_MUDA; mfg = HIJAU_TUA
    else:
        nc  = " | ".join(f"C{i+1}=({new_cen[i][0]},{new_cen[i][1]})" for i in range(3))
        msg = f"Centroid berubah — Lanjut iterasi berikutnya: {nc}"
        mbg = MERAH_MUDA; mfg = MERAH_TUA
    mg(ws,row,1,row,9,msg,bold=True,bg=mbg,fg=mfg)

def write_grafik_kmeans(wb, history, data):
    sname = 'Grafik K-Means'
    if sname in wb.sheetnames: del wb[sname]
    ws = wb.create_sheet(sname)
    ws.sheet_view.showGridLines = False
    points    = list(zip(data['Persentase'], data['Waktu']))
    soal_list = list(data['Soal'])
    ws.merge_cells('A1:N1')
    c = ws['A1']; c.value = 'VISUALISASI K-MEANS — SCATTER PLOT TIAP ITERASI'
    c.font = Font(bold=True, size=14, color=PUTIH, name='Calibri')
    c.fill = solid(BIRU_TUA); c.alignment = aln()
    ws.row_dimensions[1].height = 30
    rp = 2
    for it_num, it_data in enumerate(history, 1):
        buf = make_kmeans_scatter(it_num, it_data, points, soal_list, it_data['converged'])
        img = XLImage(buf); img.width = 750; img.height = 520
        ws.add_image(img, f'A{rp}'); rp += 29

def write_split_sheet(ws, train, test, split_detail, data_full):
    for c, w in enumerate([10,14,14,14,14,18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    row = 1
    mg(ws,row,1,row,6,"STRATIFIED SPLIT 75:25 — DATA LATIH & DATA UJI",
       bold=True, bg=BIRU_TUA, fg=PUTIH, sz=13); row += 1
    mg(ws,row,1,row,6,
       f"Total={len(data_full)} soal  |  Train={len(train)} soal (75%)  |  Test={len(test)} soal (25%)",
       bold=True, bg=BIRU_MUDA, fg=BIRU_TUA); row += 2
    mg(ws,row,1,row,6,"DISTRIBUSI PER KELAS",bold=True,bg=BIRU_MID,fg=PUTIH); row += 1
    for c, h in enumerate(["Kelas","Total","Train (75%)","Test (25%)","% Train","% Test"], 1):
        sc(ws,row,c,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row += 1
    for label in ['Mudah','Sedang','Sulit']:
        d = split_detail.get(label, {'total':0,'train':0,'test':0})
        if d['total'] == 0: continue
        k = KET.index(label)
        sc(ws,row,1,label,bold=True,bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,2,d['total'],bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,3,d['train'],bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,4,d['test'],bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,5,f"{d['train']/d['total']*100:.1f}%",bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,6,f"{d['test']/d['total']*100:.1f}%",bg=BG_C[k],fg=FG_C[k]); row += 1
    sc(ws,row,1,"TOTAL",bold=True,bg=ABU); sc(ws,row,2,len(data_full),bold=True,bg=ABU)
    sc(ws,row,3,len(train),bold=True,bg=ABU); sc(ws,row,4,len(test),bold=True,bg=ABU)
    sc(ws,row,5,"75%",bold=True,bg=ABU); sc(ws,row,6,"25%",bold=True,bg=ABU); row += 2
    mg(ws,row,1,row,6,f"DATA LATIH (TRAIN) — {len(train)} Soal",
       bold=True, bg=HIJAU_TUA, fg=PUTIH); row += 1
    for c, h in enumerate(["No","Soal","Persentase (%)","Waktu (detik)","Cluster","Keterangan"], 1):
        sc(ws,row,c,h,bold=True,bg=HIJAU_MID,fg=PUTIH)
    row += 1
    for i, r in train.iterrows():
        k = KET.index(r['Keterangan']); bg = BG_C[k]; fg = FG_C[k]
        sc(ws,row,1,i+1,bg=bg,fg=fg); sc(ws,row,2,r['Soal'],bg=bg,fg=fg)
        sc(ws,row,3,r['Persentase'],bg=bg,fg=fg); sc(ws,row,4,int(r['Waktu']),bg=bg,fg=fg)
        sc(ws,row,5,r['Cluster'],bg=bg,fg=fg); sc(ws,row,6,r['Keterangan'],bg=bg,fg=fg); row += 1
    row += 1
    mg(ws,row,1,row,6,f"DATA UJI (TEST) — {len(test)} Soal",
       bold=True, bg=MERAH_TUA, fg=PUTIH); row += 1
    for c, h in enumerate(["No","Soal","Persentase (%)","Waktu (detik)","Cluster","Keterangan"], 1):
        sc(ws,row,c,h,bold=True,bg="C00000",fg=PUTIH)
    row += 1
    for i, r in test.iterrows():
        k = KET.index(r['Keterangan']); bg = BG_C[k]; fg = FG_C[k]
        sc(ws,row,1,i+1,bg=bg,fg=fg); sc(ws,row,2,r['Soal'],bg=bg,fg=fg)
        sc(ws,row,3,r['Persentase'],bg=bg,fg=fg); sc(ws,row,4,int(r['Waktu']),bg=bg,fg=fg)
        sc(ws,row,5,r['Cluster'],bg=bg,fg=fg); sc(ws,row,6,r['Keterangan'],bg=bg,fg=fg); row += 1

def write_training_sheet(ws, model, train, classes):
    for c, w in enumerate([16,18,18,18,16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    row = 1
    mg(ws,row,1,row,5,"PROSES TRAINING — GAUSSIAN NAIVE BAYES",
       bold=True, bg=BIRU_TUA, fg=PUTIH, sz=13); row += 1
    mg(ws,row,1,row,5,
       f"Data latih: {len(train)} soal  |  Fitur: Persentase, Waktu  |  Kelas: {', '.join(classes)}",
       bold=True, bg=BIRU_MUDA, fg=BIRU_TUA); row += 2
    mg(ws,row,1,row,5,"1. PRIOR PROBABILITY  P(Ck) = jumlah_kelas / total",
       bold=True, bg=BIRU_MID, fg=PUTIH); row += 1
    for c, h in enumerate(["Kelas","Jumlah Data","Total","P(Ck)"], 1):
        sc(ws,row,c,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row += 1
    total = len(train)
    for c in classes:
        k = KET.index(c); n = list(train['Keterangan']).count(c)
        sc(ws,row,1,c,bold=True,bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,2,n,bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,3,total,bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,4,round(model.priors_[c],4),bg=BG_C[k],fg=FG_C[k]); row += 1
    row += 1
    mg(ws,row,1,row,5,"2. MEAN & STANDAR DEVIASI per KELAS",
       bold=True, bg=BIRU_MID, fg=PUTIH); row += 1
    for c, h in enumerate(["Kelas","Mean Persentase","Mean Waktu","Std Persentase","Std Waktu"], 1):
        sc(ws,row,c,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row += 1
    for c in classes:
        k = KET.index(c)
        sc(ws,row,1,c,bold=True,bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,2,round(model.means_[c][0],4),bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,3,round(model.means_[c][1],4),bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,4,round(model.stds_[c][0],4),bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,5,round(model.stds_[c][1],4),bg=BG_C[k],fg=FG_C[k]); row += 1
    row += 1
    mg(ws,row,1,row,5,
       "3. RUMUS GAUSSIAN PDF: P(x|Ck) = (1/sqrt(2pi)*sigma) x exp(-(x-mu)^2 / 2*sigma^2)",
       bold=True, bg=KUNING, fg="7F3F00", wrap=True); row += 1
    mg(ws,row,1,row,5,
       "4. LOG POSTERIOR: log P(Ck|x) = log P(Ck) + Sigma log P(xi|Ck)",
       bold=True, bg=KUNING, fg="7F3F00", wrap=True); row += 1
    mg(ws,row,1,row,5,
       "5. PREDIKSI: Kelas = argmax [ log P(Ck|x) ] (kelas dengan nilai tertinggi)",
       bold=True, bg=KUNING, fg="7F3F00")

def write_testing_sheet(ws, model, test, y_pred, y_proba, classes):
    for c, w in enumerate([8,10,14,14,14,16,16,16,14,12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    row = 1
    mg(ws,row,1,row,10,"HASIL PREDIKSI — DATA UJI (TEST)",
       bold=True, bg=BIRU_TUA, fg=PUTIH, sz=13); row += 1
    mg(ws,row,1,row,10,
       f"Data uji: {len(test)} soal  |  Metode: Log-likelihood + Softmax Probability",
       bold=True, bg=BIRU_MUDA, fg=BIRU_TUA); row += 2
    hdrs = ["No","Soal","Persentase","Waktu","Aktual",
            f"P({classes[0]})", f"P({classes[1]})", f"P({classes[2]})",
            "Prediksi","Status"]
    for c, h in enumerate(hdrs, 1):
        sc(ws,row,c,h,bold=True,bg=BIRU_TUA,fg=PUTIH,wrap=True)
    ws.row_dimensions[row].height = 28; row += 1
    for i, ((_, r), pred, prob) in enumerate(zip(test.iterrows(), y_pred, y_proba), 1):
        actual  = r['Keterangan']; correct = (pred==actual)
        k_pred  = KET.index(pred); bg = BG_C[k_pred]; fg = FG_C[k_pred]
        status  = "BENAR" if correct else "SALAH"
        sbg     = HIJAU_MUDA if correct else MERAH_MUDA
        sfg     = HIJAU_TUA  if correct else MERAH_TUA
        sc(ws,row,1,i,bg=bg,fg=fg); sc(ws,row,2,r['Soal'],bg=bg,fg=fg)
        sc(ws,row,3,r['Persentase'],bg=bg,fg=fg); sc(ws,row,4,int(r['Waktu']),bg=bg,fg=fg)
        sc(ws,row,5,actual,bold=True,bg=bg,fg=fg)
        for ci, p in enumerate(prob, 6):
            sc(ws,row,ci,f"{p*100:.2f}%",bg=bg,fg=fg)
        sc(ws,row,9,pred,bold=True,bg=bg,fg=fg)
        sc(ws,row,10,status,bold=True,bg=sbg,fg=sfg); row += 1

def write_evaluasi_sheet(ws, y_true, y_pred, cm_df, met_df, acc, classes,
                         buf_cm, buf_bar, buf_scatter):
    for c, w in enumerate([16,14,14,14,14,14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    row = 1
    mg(ws,row,1,row,6,"EVALUASI MODEL — GAUSSIAN NAIVE BAYES",
       bold=True, bg=BIRU_TUA, fg=PUTIH, sz=13); row += 1
    mg(ws,row,1,row,6,
       f"Akurasi: {acc*100:.2f}%  |  Data uji: {len(y_true)} soal  |  Kelas: {', '.join(classes)}",
       bold=True, bg=HIJAU_MUDA, fg=HIJAU_TUA, sz=12); row += 2
    mg(ws,row,1,row,6,"1. AKURASI",bold=True,bg=BIRU_MID,fg=PUTIH); row += 1
    benar = sum(1 for t,p in zip(y_true,y_pred) if t==p)
    sc(ws,row,1,"Prediksi Benar",bold=True,bg=HIJAU_MUDA,fg=HIJAU_TUA)
    sc(ws,row,2,benar,bg=HIJAU_MUDA,fg=HIJAU_TUA)
    sc(ws,row,3,"Total Data Uji",bold=True,bg=ABU); sc(ws,row,4,len(y_true),bg=ABU)
    sc(ws,row,5,"Akurasi",bold=True,bg=KUNING,fg="7F3F00")
    sc(ws,row,6,f"{acc*100:.2f}%",bold=True,bg=KUNING,fg="7F3F00"); row += 2
    mg(ws,row,1,row,6,"2. CONFUSION MATRIX  (Baris=Aktual, Kolom=Prediksi)",
       bold=True, bg=BIRU_MID, fg=PUTIH); row += 1
    sc(ws,row,1,"Aktual \\ Prediksi",bold=True,bg=BIRU_TUA,fg=PUTIH)
    for ci, c in enumerate(classes, 2):
        k = KET.index(c); sc(ws,row,ci,c,bold=True,bg=BG_C[k],fg=FG_C[k])
    row += 1
    for ri, ra in enumerate(classes):
        kr = KET.index(ra); sc(ws,row,1,ra,bold=True,bg=BG_C[kr],fg=FG_C[kr])
        for ci, ca in enumerate(classes, 2):
            val = int(cm_df.loc[ra,ca])
            bg  = HIJAU_MUDA if ra==ca else MERAH_MUDA
            fg  = HIJAU_TUA  if ra==ca else MERAH_TUA
            sc(ws,row,ci,val,bold=(ra==ca),bg=bg,fg=fg); row += 1
    row += 1
    mg(ws,row,1,row,6,"3. PRECISION / RECALL / F1-SCORE",bold=True,bg=BIRU_MID,fg=PUTIH); row += 1
    for c, h in enumerate(["Kelas","Precision","Recall","F1-Score","Support"], 1):
        sc(ws,row,c,h,bold=True,bg=BIRU_TUA,fg=PUTIH)
    row += 1
    supports = {c: list(y_true).count(c) for c in classes}
    for c in classes:
        k = KET.index(c)
        sc(ws,row,1,c,bold=True,bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,2,f"{met_df.loc[c,'Precision']:.4f}",bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,3,f"{met_df.loc[c,'Recall']:.4f}",   bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,4,f"{met_df.loc[c,'F1']:.4f}",       bg=BG_C[k],fg=FG_C[k])
        sc(ws,row,5,supports[c],                        bg=BG_C[k],fg=FG_C[k]); row += 1
    sc(ws,row,1,"Macro Avg",bold=True,bg=ABU)
    sc(ws,row,2,f"{met_df['Precision'].mean():.4f}",bold=True,bg=ABU)
    sc(ws,row,3,f"{met_df['Recall'].mean():.4f}",   bold=True,bg=ABU)
    sc(ws,row,4,f"{met_df['F1'].mean():.4f}",       bold=True,bg=ABU)
    sc(ws,row,5,len(y_true),bold=True,bg=ABU); row += 2
    mg(ws,row,1,row,6,"4. VISUALISASI",bold=True,bg=BIRU_MID,fg=PUTIH); row += 1
    img_cm  = XLImage(buf_cm);     img_cm.width=380;  img_cm.height=310
    img_bar = XLImage(buf_bar);    img_bar.width=510; img_bar.height=330
    img_sc  = XLImage(buf_scatter);img_sc.width=600;  img_sc.height=415
    ws.add_image(img_cm,  f'A{row}')
    ws.add_image(img_bar, f'F{row}')
    row += 18
    ws.add_image(img_sc, f'A{row}')

def write_grafik_nb_sheet(wb, model, train, test, y_pred, y_proba, data, classes):
    sname = 'Grafik NB'
    if sname in wb.sheetnames: del wb[sname]
    ws = wb.create_sheet(sname)
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:N1')
    c = ws['A1']; c.value = 'VISUALISASI GAUSSIAN NAIVE BAYES — GRAFIK LENGKAP'
    c.font = Font(bold=True, size=14, color=PUTIH, name='Calibri')
    c.fill = solid(BIRU_TUA); c.alignment = aln()
    ws.row_dimensions[1].height = 30
    rp = 2
    ws.merge_cells(f'A{rp}:N{rp}')
    cc = ws[f'A{rp}']; cc.value = '1. KURVA DISTRIBUSI GAUSSIAN PDF'
    cc.font = Font(bold=True,size=12,color=BIRU_TUA,name='Calibri')
    cc.fill = solid(BIRU_MUDA); cc.alignment = aln(); rp += 1
    img1 = XLImage(make_gaussian_pdf(model, train, classes))
    img1.width = 860; img1.height = 365
    ws.add_image(img1, f'A{rp}'); rp += 21
    ws.merge_cells(f'A{rp}:N{rp}')
    cc = ws[f'A{rp}']; cc.value = '2. PROBABILITAS PREDIKSI TIAP SOAL UJI'
    cc.font = Font(bold=True,size=12,color=BIRU_TUA,name='Calibri')
    cc.fill = solid(BIRU_MUDA); cc.alignment = aln(); rp += 1
    img2 = XLImage(make_proba_bar(test, y_pred, y_proba, classes))
    img2.width = 720; img2.height = 400
    ws.add_image(img2, f'A{rp}'); rp += 22
    ws.merge_cells(f'A{rp}:N{rp}')
    cc = ws[f'A{rp}']; cc.value = '3. DISTRIBUSI HASIL CLUSTERING — PROPORSI PER KELAS'
    cc.font = Font(bold=True,size=12,color=BIRU_TUA,name='Calibri')
    cc.fill = solid(BIRU_MUDA); cc.alignment = aln(); rp += 1
    img3 = XLImage(make_pie_cluster(data, classes))
    img3.width = 800; img3.height = 370
    ws.add_image(img3, f'A{rp}')


# ─────────────────────────────────────────────
# MAIN GENERATE EXCEL
# ─────────────────────────────────────────────
def generate_excel(df_raw, data, history, train, test, split_detail,
                   model, y_pred, y_proba, y_test, classes,
                   cm_df, met_df, acc, info):
    buf_cm      = make_cm_heatmap(cm_df, classes)
    buf_bar     = make_metrics_bar(met_df, acc)
    buf_scatter = make_scatter_nb(train, test, y_pred, classes)

    wb = Workbook()

    # Sheet Data Asli
    ws_data = wb.active; ws_data.title = "Data"
    write_sheet_data_asli(ws_data, df_raw)

    # Sheet Persentase
    write_sheet_persentase(
        wb.create_sheet("Persentase"), data,
        info['has_jumlah_benar'], info['has_total_siswa']
    )

    # Sheet Waktu
    write_sheet_waktu(wb.create_sheet("Waktu"), data)

    # Sheet Iterasi K-Means
    prev = None
    for it_num, it_data in enumerate(history, 1):
        sname = f"Iterasi {it_num} (Final)" if it_data['converged'] else f"Iterasi {it_num}"
        write_iterasi_sheet(wb.create_sheet(sname), it_num, it_data, data, prev)
        prev = it_data['assignments'][:]

    # Sheet NB
    write_split_sheet(wb.create_sheet("Stratified Split"),
                      train, test, split_detail, data)
    write_training_sheet(wb.create_sheet("Training NB"), model, train, classes)
    write_testing_sheet(wb.create_sheet("Testing NB"),
                        model, test, y_pred, y_proba, classes)
    write_evaluasi_sheet(wb.create_sheet("Evaluasi NB"),
                         y_test, y_pred, cm_df, met_df, acc, classes,
                         buf_cm, buf_bar, buf_scatter)

    # Sheet Grafik NB
    write_grafik_nb_sheet(wb, model, train, test, y_pred, y_proba, data, classes)

    # Sheet Grafik K-Means (pindah ke depan)
    write_grafik_kmeans(wb, history, data)
    wb.move_sheet('Grafik K-Means', offset=-len(wb.sheetnames)+1)
    wb.move_sheet('Grafik NB',      offset=-len(wb.sheetnames)+2)

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out


# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────
st.title("🔬 Pipeline K-Means + Gaussian Naive Bayes")
st.markdown("Klasifikasi tingkat kesulitan soal — **Excel lengkap dengan warna, grafik, dan semua sheet**")

with st.sidebar:
    st.header("⚙️ Pengaturan")
    test_ratio = st.slider("Rasio Data Test", 0.10, 0.40, 0.25, 0.05,
                           help="Default 25% untuk test")
    st.caption("Versi 2.0 — Pipeline Manual Lengkap")
    st.markdown("---")
    st.markdown("""
**Format kolom yang diterima:**
- `Soal` / `No` / `Nomor Soal`
- `Persentase` / `Persen` ← **ATAU**
- `Jumlah Benar` + `Total Siswa`
- `Waktu` / `Waktu (detik)`
""")

uploaded_file = st.file_uploader(
    "📂 Upload File Excel (.xlsx)",
    type=["xlsx"],
    help="File harus memiliki kolom Soal, Persentase/Jumlah Benar, dan Waktu"
)

if uploaded_file is not None:
    try:
        df_raw = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Gagal membaca file: {e}"); st.stop()

    st.success(f"✅ File dimuat: **{len(df_raw)} baris**, {len(df_raw.columns)} kolom")

    with st.expander("👁️ Preview Data Mentah"):
        st.dataframe(df_raw.head(10), use_container_width=True)

    # ── Deteksi & Persiapan Data ──
    try:
        data, info = load_and_prepare(df_raw)
    except ValueError as e:
        st.error(f"❌ {e}"); st.stop()

    col1, col2, col3 = st.columns(3)
    col1.metric("Total Soal", len(data))
    col2.metric("Deteksi Jumlah Benar", "✅" if info['has_jumlah_benar'] else "—")
    col3.metric("Deteksi Total Siswa",  "✅" if info['has_total_siswa']  else "—")

    if info['has_jumlah_benar'] and info['has_total_siswa']:
        st.info("📊 Persentase dihitung otomatis dari **Jumlah Benar / Total Siswa × 100**")

    with st.expander("👁️ Preview Data yang Diproses"):
        st.dataframe(data, use_container_width=True)

    st.markdown("---")
    st.subheader("🚀 Jalankan Pipeline")

    if st.button("▶️ Proses & Generate Excel Lengkap", type="primary", use_container_width=True):
        progress = st.progress(0, text="Memulai pipeline...")

        with st.spinner("⏳ K-Means Clustering..."):
            history = run_kmeans(data)
            final   = history[-1]
            data['Cluster']    = [f"C{a+1}" for a in final['assignments']]
            data['Keterangan'] = [KET[a]     for a in final['assignments']]
        progress.progress(25, text="K-Means selesai...")

        with st.spinner("⏳ Stratified Split..."):
            train, test, split_detail = stratified_split(data, test_ratio=test_ratio)
        progress.progress(40, text="Split selesai...")

        with st.spinner("⏳ Training Gaussian Naive Bayes..."):
            classes = sorted(data['Keterangan'].unique(), key=lambda x: KET.index(x))
            X_train = train[['Persentase','Waktu']].values
            y_train = train['Keterangan'].values
            model   = GaussianNaiveBayes()
            model.fit(X_train, y_train)
        progress.progress(55, text="Training selesai...")

        with st.spinner("⏳ Prediksi & Evaluasi..."):
            X_test  = test[['Persentase','Waktu']].values
            y_test  = test['Keterangan'].values
            y_pred  = model.predict(X_test)
            y_proba = model.predict_proba(X_test)
            acc     = np.mean(y_test == y_pred)
            cm_df   = cm_manual(y_test, y_pred, classes)
            met_df  = metrics_manual(cm_df)
        progress.progress(70, text="Evaluasi selesai...")

        with st.spinner("⏳ Generate Excel lengkap (grafik + semua sheet)..."):
            uploaded_file.seek(0)
            excel_buf = generate_excel(
                df_raw, data, history, train, test, split_detail,
                model, y_pred, y_proba, y_test, classes,
                cm_df, met_df, acc, info
            )
        progress.progress(95, text="Excel selesai...")

        model_buf = model.save_bytes()
        progress.progress(100, text="✅ Selesai!")
        st.balloons()

        # ── Hasil di UI ──
        st.markdown("---")
        st.subheader("📊 Hasil Pipeline")
        tabs = st.tabs(["K-Means", "Naive Bayes", "Visualisasi"])

        with tabs[0]:
            st.write(f"**Konvergen dalam {len(history)} iterasi**")
            st.dataframe(data[['Soal','Persentase','Waktu','Cluster','Keterangan']],
                         use_container_width=True)
            dist = data['Keterangan'].value_counts().reset_index()
            dist.columns = ['Kelas','Jumlah']
            st.bar_chart(dist.set_index('Kelas'))

        with tabs[1]:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Akurasi",   f"{acc*100:.2f}%")
            c2.metric("Macro P",   f"{met_df['Precision'].mean():.4f}")
            c3.metric("Macro R",   f"{met_df['Recall'].mean():.4f}")
            c4.metric("Macro F1",  f"{met_df['F1'].mean():.4f}")
            st.write("**Confusion Matrix**")
            st.dataframe(cm_df, use_container_width=True)
            st.write("**Metrics per Kelas**")
            st.dataframe(met_df.round(4), use_container_width=True)

        with tabs[2]:
            fig = px.scatter(data, x='Persentase', y='Waktu', color='Keterangan',
                             symbol='Keterangan', title="Hasil Clustering K-Means",
                             color_discrete_map={'Mudah':'#2E75B6','Sedang':'#70AD47','Sulit':'#E00000'})
            st.plotly_chart(fig, use_container_width=True)

        # ── Download ──
        st.markdown("---")
        st.subheader("📥 Download Hasil")
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.download_button(
                label="📥 Download Excel Lengkap",
                data=excel_buf,
                file_name=f"Hasil_KMeans_NB_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
        with col_d2:
            st.download_button(
                label="📥 Download Model (.pkl)",
                data=model_buf,
                file_name="model_nb.pkl",
                use_container_width=True
            )

        # ── Info Sheet ──
        with st.expander("📋 Daftar Sheet Excel yang dibuat"):
            sheets = (
                ["Data", "Persentase", "Waktu"] +
                [f"Iterasi {i+1}" + (" (Final)" if it['converged'] else "")
                 for i, it in enumerate(history)] +
                ["Stratified Split","Training NB","Testing NB","Evaluasi NB",
                 "Grafik NB","Grafik K-Means"]
            )
            for s in sheets:
                st.markdown(f"- `{s}`")

else:
    st.info("⬆️ Silakan upload file Excel berisi data soal.")
    st.markdown("""
### Contoh format kolom yang diterima:

| Soal | Jumlah Benar | Total Siswa | Waktu (detik) |
|------|-------------|-------------|---------------|
| 1    | 18          | 22          | 45            |
| 2    | 15          | 22          | 62            |

**atau dengan kolom Persentase langsung:**

| No Soal | Persentase | Waktu |
|---------|-----------|-------|
| S1      | 81.82     | 45    |
""")

st.caption("Pipeline K-Means Manual + Gaussian Naive Bayes Manual • Excel identik dengan skrip standalone")
